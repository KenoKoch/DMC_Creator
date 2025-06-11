import tkinter as tk
from tkinter import ttk, font, messagebox
from pylibdmtx.pylibdmtx import encode
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
import os
import xlwings as xw



#Variablen

def get_inputs():
    ergebnisse = {} 
    # Fügen Ergebnisse hinzu
    ergebnisse['Station'] = auswahlmenu_var.get()  # Stationsauswahl
    ergebnisse['SL'] = (entry_SL_NR.get(), entry_SL_DT.get())  # SL
    ergebnisse['STB'] = (entry_STB_NR.get(), entry_STB_DT.get())  # STB
    ergebnisse['STW'] = (entry_STW_NR.get(), entry_STW_DT.get())  # STW
    ergebnisse['SmartHead'] = (entry_SH_NR.get(), entry_SH_DT.get())  # SH
    ergebnisse['WeraGrün'] = (entry_WG_NR.get(), entry_WG_DT.get())  # Wera Grün
    ergebnisse['WeraGelb'] = (entry_WGe_NR.get(), entry_WGe_DT.get())  # Wera Gelb
    ergebnisse['Multimeter'] = (entry_MT_NR.get(), entry_MT_DT.get())  # Multimeter
    ergebnisse['Metriso'] = (entry_MI_NR.get(), entry_MI_DT.get())  # Metriso
    ergebnisse['Microohmmeter'] = (entry_MO_NR.get(), entry_MO_DT.get())  # Microohmmeter

    return ergebnisse

# Generien der Excel sheets wenn nicht vorhanden
def generate_dmc_code():
    Inputs = get_inputs()
    wb = xw.Book()
    ws = wb.sheets.active
    Bilder_loeschen = []
    Erfolgreich = False
    
    if Inputs['Station'] != "Prüfgeräte":
        row = 3
        col = 'A'
        if all(value not in ["",''] for key in ['SL','STB','STW','SmartHead','WeraGrün','WeraGelb'] for value in Inputs.get(key)):
            for geraet, wert in Inputs.items():
                if geraet in ['SL', 'STB', 'STW', 'SmartHead','WeraGrün','WeraGelb'] and isinstance(wert, tuple) and len(wert) == 2:
                        nr, dt = wert
                        DMC_IN = "Geraet: " + nr + " Pruefdatum: " + dt
                        encoded = encode(DMC_IN.encode())
                        DMC = Image.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
                        output = Inputs["Station"]+"_DMC_" + geraet + ".png"
                        DMC.save(output)
                        print(os.path.exists(output))
                        ws.pictures.add(output, name=geraet, left=ws.range(col + str(row + 1)).left, top=ws.range(col + str(row + 1)).top)
                        Erfolgreich = True
                        Bilder_loeschen.append(output)
                        if col == 'A':
                            col = 'E'
                        else:
                            col = 'A'
                            row += 10
        else:
            messagebox.showwarning("Eingabe leer !)", "Bitte füllen Sie alle Felder aus")
                    
    else:
        row = 3
        col = 'A'
        if all(value not in ["",''] for key in ['Multimeter', 'Metriso', 'Microohmmeter'] for value in Inputs.get(key)):
            for geraet, wert in Inputs.items():
                if geraet in ['Multimeter', 'Metriso', 'Microohmmeter'] and isinstance(wert, tuple) and len(wert) == 2:
                    nr, dt = wert
                    DMC_IN = "Geraet: " + nr + " Pruefdatum: " + dt
                    encoded = encode(DMC_IN.encode())
                    DMC = Image.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
                    output = Inputs["Station"]+"_DMC_" + geraet + ".png"
                    DMC.save(output)
                    img = ExcelImage(output)
                    ws[col + str(row)] = geraet
                    ws.add_image(img, col + str(row + 1))
                    Erfolgreich = True
                    Bilder_loeschen.append(output)
                    if col == 'A':
                        col = 'E'
                    else:
                        col = 'A'
                        row += 10
        else:
            messagebox.showwarning("Eingabe leer !)", "Bitte füllen Sie alle Felder aus")
    
    if Erfolgreich :
        ws.cell(row=1, column=3, value=Inputs["Station"]).font=Font(size=20)
        Dateiname= 'DMC_'+ Inputs["Station"] + '.xlsx'
        wb.save(Dateiname)
        for Bilder in Bilder_loeschen:
            os.remove(Bilder)
        messagebox.showinfo("Generiert :)","Excel erfolgreich generiert")
    
def edit_DMC_Code():
    Inputs = get_inputs()
    Bilder_loeschen = []
    wb = None

    if Inputs['Station'] != "Prüfgeräte":
        Erfolgreich = False
        for geraet, wert in Inputs.items():
            if geraet in ['SL', 'STB', 'STW', 'SmartHead','WeraGrün','WeraGelb'] and isinstance(wert, tuple) and len(wert) == 2 and any(wert):
                nr, dt = wert
                if nr and dt :
                    DMC_IN = "Geraet: " + nr + " Pruefdatum: " + dt
                    encoded = encode(DMC_IN.encode())
                    DMC = Image.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
                    output = Inputs["Station"]+"_DMC_" + geraet + ".png"
                    DMC.save(output)
                    img = ExcelImage(output)
                    wb = load_workbook('DMC_'+Inputs["Station"]+'.xlsx')
                    ws = wb.active
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value == geraet:
                                new_coordinate = ws[cell.coordinate[0] + str(int(cell.coordinate[1:]) + 1)] # Zeile 1 unter Name
                                ws.add_image(img, new_coordinate.coordinate) # Füge das Bild hinzu
                    wb.save('DMC_'+Inputs["Station"]+'.xlsx')
                    Erfolgreich = True
                    Bilder_loeschen.append(output)
                else:
                    messagebox.showwarning("Eingabe leer !)","Bitte füllen Sie Feld Nummer und Prüfdatum aus")   
    else:
        Erfolgreich = False
        for geraet, wert in Inputs.items():
            if geraet in ['Multimeter', 'Metriso', 'Microohmmeter'] and isinstance(wert, tuple) and len(wert) == 2 and any(wert):
                nr, dt = wert
                if nr and dt :
                    DMC_IN = "Geraet: " + nr + " Pruefdatum: " + dt
                    encoded = encode(DMC_IN.encode())
                    DMC = Image.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
                    output = Inputs["Station"]+"_DMC_" + geraet + ".png"
                    DMC.save(output)
                    img = ExcelImage(output)
                    wb = load_workbook('DMC_'+Inputs["Station"]+'.xlsx')
                    ws = wb.active
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value == geraet:
                                new_coordinate = ws[cell.coordinate[0] + str(int(cell.coordinate[1:]) + 1)] # Speichere das Bild in der nächsten Zeile
                                ws.add_image(img, new_coordinate.coordinate) # Fügen das Bild hinzu
                    wb.save('DMC_'+Inputs["Station"]+'.xlsx')
                    Erfolgreich = True
                    Bilder_loeschen.append(output)
                else:
                    messagebox.showwarning("Eingabe leer !)","Bitte füllen Sie Feld Nummer und Prüfdatum aus")

    for Bilder in Bilder_loeschen:
        os.remove(Bilder)
    if Erfolgreich:
        messagebox.showinfo("Editiert :)","Excel erfolgreich editiert")


# User Interface
root = tk.Tk()
style = ttk.Style()

#Schriftart
label1_font = font.Font(family="Arial", size=18, weight="bold", underline=True)
label_font = ("Arial", 14)
button_font = ("Arial", 12, "bold")
entry_font = ("Arial", 12)

# Überschrift
label1 = ttk.Label(root, text="DMC Code Generator", font=label1_font, underline=True)
label1.grid(row=0, column=0, columnspan=2, pady=(0,10), padx=20)

# Auswahlmenü
optionen = ["Station auswählen","Station 1", "Station 2", "Station 3", "Station 4","Prüfgeräte"]
auswahlmenu_var = tk.StringVar()
auswahlmenu_var.set(optionen[0]) 

label2 = ttk.Label(root, text="Station:", font=label_font)
label2.grid(row=1, column=0, columnspan=2)

auswahlmenu = ttk.OptionMenu(root, auswahlmenu_var, *optionen, style="TMenubutton")
auswahlmenu.grid(row=2, column=0, columnspan=2, pady=(0,5))

style.configure("TMenubutton", font=entry_font, width=17)

# Eingabe SL Schrauber
label_SL = ttk.Label(root, text="SL-Schrauber:", font=label_font)
label_SL_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_SL_NR = ttk.Entry(root, font=entry_font)
label_SL_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_SL_DT = ttk.Entry(root, font=entry_font)

# Eingabe STB Schrauber
label_STB = ttk.Label(root, text="STB-Schrauber:", font=label_font)
label_STB_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_STB_NR = ttk.Entry(root, font=entry_font)
label_STB_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_STB_DT = ttk.Entry(root, font=entry_font)

# Eingabe ST-Wrench
label_STW = ttk.Label(root, text="ST-Wrench:", font=label_font)
label_STW_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_STW_NR = ttk.Entry(root,font=entry_font)
label_STW_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_STW_DT= ttk.Entry(root,font=entry_font)

# Eingabe Smart Head
label_SH = ttk.Label(root, text="Smart Head:", font=label_font)
label_SH_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_SH_NR = ttk.Entry(root,font=entry_font)
label_SH_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_SH_DT= ttk.Entry(root,font=entry_font)

# Eingabe Wera Grün
label_WG = ttk.Label(root, text="Wera Grün:", font=label_font)
label_WG_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_WG_NR = ttk.Entry(root,font=entry_font)
label_WG_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_WG_DT= ttk.Entry(root,font=entry_font)

# Eingabe Wera Gelb
label_WGe = ttk.Label(root, text="Wera Gelb:", font=label_font)
label_WGe_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_WGe_NR = ttk.Entry(root,font=entry_font)
label_WGe_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_WGe_DT= ttk.Entry(root,font=entry_font)

# Eingabe Prüfgeräte

# Eingabe Multimeter
label_MT = ttk.Label(root, text="Multimeter:", font=label_font)
label_MT_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_MT_NR = ttk.Entry(root, font=entry_font)
label_MT_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_MT_DT = ttk.Entry(root, font=entry_font)

# Eingabe Metriso
label_MI = ttk.Label(root, text="Metriso:", font=label_font)
label_MI_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_MI_NR = ttk.Entry(root, font=entry_font)
label_MI_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_MI_DT = ttk.Entry(root, font=entry_font)

# Eingabe Micoohmmeter
label_MO = ttk.Label(root, text="Microohmmeter:", font=label_font)
label_MO_NR = ttk.Label(root, text="Nummer: ", font=entry_font)
entry_MO_NR = ttk.Entry(root,font=entry_font)
label_MO_DT = ttk.Label(root, text="Prüfdatum: ", font=entry_font)
entry_MO_DT= ttk.Entry(root,font=entry_font)

# Button DMC generieren
button = ttk.Button(root, text="Generiere Excel", command=generate_dmc_code, style="Primary.TButton", width=20)
style.configure("Primary.TButton", font=button_font)

# Button Excel editieren
button_2 = ttk.Button(root, text="Editiere Excel", command=edit_DMC_Code, style="Primary.TButton", width=20)
style.configure("Primary.TButton", font=button_font)

def update_visibility(*args):
    if auswahlmenu_var.get() == "Prüfgeräte":
        label_SL.grid_remove()
        label_SL_NR.grid_remove()
        entry_SL_NR.grid_remove()
        label_SL_DT.grid_remove()
        entry_SL_DT.grid_remove()

        label_STB.grid_remove()
        label_STB_NR.grid_remove()
        entry_STB_NR.grid_remove()
        label_STB_DT.grid_remove()
        entry_STB_DT.grid_remove()

        label_STW.grid_remove()
        label_STW_NR.grid_remove()
        entry_STW_NR.grid_remove()
        label_STW_DT.grid_remove()
        entry_STW_DT.grid_remove()

        label_SH.grid_remove()
        label_SH_NR.grid_remove()
        entry_SH_NR.grid_remove()
        label_SH_DT.grid_remove()
        entry_SH_DT.grid_remove()

        label_WG.grid_remove()
        label_WG_NR.grid_remove()
        entry_WG_NR.grid_remove()
        label_WG_DT.grid_remove()
        entry_WG_DT.grid_remove()

        label_WGe.grid_remove()
        label_WGe_NR.grid_remove()
        entry_WGe_NR.grid_remove()
        label_WGe_DT.grid_remove()
        entry_WGe_DT.grid_remove()

        label_MT.grid(row=3, column=0,columnspan=2)
        label_MT_NR.grid(row=4, column=0, sticky="e")
        entry_MT_NR.grid(row=4, column=1)
        label_MT_DT.grid(row=5, column=0, sticky="e")
        entry_MT_DT.grid(row=5, column=1,pady=(1,10))

        label_MI.grid(row=6, column=0, columnspan=2)
        label_MI_NR.grid(row=7, column=0, sticky="e")
        entry_MI_NR.grid(row=7, column=1)
        label_MI_DT.grid(row=8, column=0, sticky="e")
        entry_MI_DT.grid(row=8, column=1,pady=(1,10))

        label_MO.grid(row=9, column=0, columnspan=2)
        label_MO_NR.grid(row=10, column=0, sticky="e")
        entry_MO_NR.grid(row=10, column=1)
        label_MO_DT.grid(row=11, column=0, sticky="e")
        entry_MO_DT.grid(row=11, column=1, pady=(1,10))

        button.grid(row=13, column=0, columnspan=2, pady=(0,2),)
        button_2.grid(row=14, column=0, columnspan=2, pady=(0,5),)

    else:
        label_SL.grid(row=3, column=0,columnspan=2)
        label_SL_NR.grid(row=4, column=0, sticky="e")
        entry_SL_NR.grid(row=4, column=1)
        label_SL_DT.grid(row=5, column=0, sticky="e")
        entry_SL_DT.grid(row=5, column=1,pady=(1,10))

        label_STB.grid(row=6, column=0, columnspan=2)
        label_STB_NR.grid(row=7, column=0, sticky="e")
        entry_STB_NR.grid(row=7, column=1)
        label_STB_DT.grid(row=8, column=0, sticky="e")
        entry_STB_DT.grid(row=8, column=1,pady=(1,10))

        label_STW.grid(row=9, column=0, columnspan=2)
        label_STW_NR.grid(row=10, column=0, sticky="e")
        entry_STW_NR.grid(row=10, column=1,)
        label_STW_DT.grid(row=11, column=0, sticky="e")
        entry_STW_DT.grid(row=11, column=1, pady=(1,10))

        label_SH.grid(row=12, column=0, columnspan=2)
        label_SH_NR.grid(row=13, column=0, sticky="e")
        entry_SH_NR.grid(row=13, column=1)
        label_SH_DT.grid(row=14, column=0, sticky="e")
        entry_SH_DT.grid(row=14, column=1, pady=(1,10))

        label_WG.grid(row=15, column=0, columnspan=2)
        label_WG_NR.grid(row=16, column=0, sticky="e")
        entry_WG_NR.grid(row=16, column=1)
        label_WG_DT.grid(row=17, column=0, sticky="e")
        entry_WG_DT.grid(row=17, column=1, pady=(1,10))

        label_WGe.grid(row=18, column=0, columnspan=2)
        label_WGe_NR.grid(row=19, column=0, sticky="e")
        entry_WGe_NR.grid(row=19, column=1)
        label_WGe_DT.grid(row=20, column=0, sticky="e")
        entry_WGe_DT.grid(row=20, column=1, pady=(1,10))

        button.grid(row=21, column=0, columnspan=2, pady=(0,2))
        button_2.grid(row=22, column=0, columnspan=2, pady=(0,5))

        label_MT.grid_remove()
        label_MT_NR.grid_remove()
        entry_MT_NR.grid_remove()
        label_MT_DT.grid_remove()
        entry_MT_DT.grid_remove()

        label_MI.grid_remove()
        label_MI_NR.grid_remove()
        entry_MI_NR.grid_remove()
        label_MI_DT.grid_remove()
        entry_MI_DT.grid_remove()

        label_MO.grid_remove()
        label_MO_NR.grid_remove()
        entry_MO_NR.grid_remove()
        label_MO_DT.grid_remove()
        entry_MO_DT.grid_remove()


auswahlmenu_var.trace_add('write', update_visibility)

root.mainloop()



